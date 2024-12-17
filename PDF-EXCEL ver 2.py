import pdfplumber
import openpyxl
import os
from PIL import Image
import pytesseract

# Define paths
pdf_path = "path_to_your_pdf.pdf"  # Replace with your PDF path
excel_file = "output_excel_file.xlsx"  # Replace with desired output Excel file path

# Check if PDF exists
if not os.path.exists(pdf_path):
    print(f"PDF file not found: {pdf_path}")
else:
    # Open the PDF with pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        default_sheet.title = "Summary"

        # Loop through each page and extract table data
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"Processing page {page_num}...")

            # Extract tables
            tables = page.extract_tables()

            if tables:
                # Write tables to Excel (each table on a new sheet)
                for table_index, table in enumerate(tables):
                    # Create a new sheet for each table
                    sheet_name = f"Page{page_num}_Table{table_index + 1}"
                    new_sheet = workbook.create_sheet(sheet_name)

                    # Write the table rows and columns to the new sheet
                    for row_idx, row in enumerate(table, start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            new_sheet.cell(row=row_idx, column=col_idx, value=cell)
            else:
                print(f"No tables found on page {page_num}.")

        # Remove the default sheet if unused
        if not default_sheet.max_row > 1:
            workbook.remove(default_sheet)

        # Save the Excel file
        workbook.save(excel_file)
        print(f"Excel file saved: {excel_file}")
