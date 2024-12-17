import pdfplumber
import openpyxl
import os
from PIL import Image
import pytesseract

# Define paths
pdf_path = "path_to_your_pdf.pdf"  # Replace with your PDF file path
excel_file = "output_excel_file.xlsx"  # Replace with your desired output file path

# Check if PDF exists
if not os.path.exists(pdf_path):
    print(f"PDF file not found: {pdf_path}")
else:
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Summary"

    # Open the PDF with pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        # Loop through each page and process it
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"Processing page {page_num}...")

            # Extract tables from the page
            tables = page.extract_tables()
            if tables:
                print(f"Tables found on page {page_num}. Writing to Excel...")
                # Write tables to Excel
                for table_index, table in enumerate(tables):
                    # Create a new sheet for each table
                    sheet_name = f"Page{page_num}_Table{table_index + 1}"
                    new_sheet = workbook.create_sheet(title=sheet_name)

                    # Write rows and columns of the table
                    for row_num, row in enumerate(table, start=1):
                        for col_num, cell in enumerate(row, start=1):
                            new_sheet.cell(row=row_num, column=col_num, value=cell)
            else:
                print(f"No tables found on page {page_num}. Attempting OCR...")

                # Convert PDF page to image for OCR
                pil_image = page.to_image(resolution=300).original  # Use higher resolution for better OCR
                image_path = f"page_{page_num}.png"
                pil_image.save(image_path)

                # Perform OCR on the saved image
                text = pytesseract.image_to_string(Image.open(image_path), lang="eng")
                print(f"OCR text from page {page_num}: {text[:200]}...")  # Debug first 200 characters

                # Write OCR text to a separate sheet
                ocr_sheet_name = f"Page{page_num}_OCR"
                ocr_sheet = workbook.create_sheet(title=ocr_sheet_name)

                # Write OCR text (line by line) to Excel
                for row_num, line in enumerate(text.split("\n"), start=1):
                    ocr_sheet.cell(row=row_num, column=1, value=line)

                # Clean up the temporary image file
                os.remove(image_path)

    # Remove default sheet if unused
    if not default_sheet.max_row > 1:
        workbook.remove(default_sheet)

    # Save the Excel file
    workbook.save(excel_file)
    print(f"Excel file saved at: {excel_file}")
