import pdfplumber
import openpyxl
import os
from PIL import Image
import pytesseract
import cv2
import numpy as np

# Define paths
pdf_path = r"C:\Users\jamai\OneDrive\Desktop\FOR CODING\BIR FORMS\form.pdf"
excel_file = r"C:\Users\jamai\OneDrive\Desktop\FOR CODING\BIR FORMS\form.xlsx"

# Check if PDF exists
if not os.path.exists(pdf_path):
    print(f"PDF file not found: {pdf_path}")
else:
    # Open the PDF with pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        # Create a new Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Loop through each page and extract table data
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"Processing page {page_num}...")

            # Attempt to extract tables
            tables = page.extract_tables()
            print(f"Tables found on page {page_num}: {tables}")  # Debugging line

            if tables:
                # Write tables to Excel
                for table_index, table in enumerate(tables):
                    sheet_name = f"Page {page_num}_Table {table_index + 1}"
                    new_sheet = workbook.create_sheet(sheet_name)

                    # Write rows of the table to the new sheet
                    for row_num, row in enumerate(table, start=1):
                        for col_num, cell in enumerate(row, start=1):
                            new_sheet.cell(row=row_num, column=col_num, value=cell)
                print(f"Table(s) from page {page_num} written to Excel.")
            else:
                # Attempt OCR if no tables are found
                print(f"No tables found on page {page_num}. Attempting OCR...")

                # Convert PDF page to image
                pil_image = page.to_image()
                image_path = f"page_{page_num}.png"
                pil_image.save(image_path)

                # Preprocess image for better OCR accuracy
                image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                _, thresh_image = cv2.threshold(image, 150, 255, cv2.THRESH_BINARY)

                # Perform OCR on the preprocessed image
                text = pytesseract.image_to_string(thresh_image)
                print(f"OCR text from page {page_num}: {text[:200]}...")  # Debugging OCR output

                # Write OCR text to Excel
                ocr_sheet_name = f"Page_{page_num}_OCR"
                new_sheet = workbook.create_sheet(ocr_sheet_name)
                new_sheet.cell(row=1, column=1, value=text)
                print(f"OCR text from page {page_num} written to Excel.")

        # Save the Excel file
        workbook.save(excel_file)
        print(f"Excel file saved at: {excel_file}")
